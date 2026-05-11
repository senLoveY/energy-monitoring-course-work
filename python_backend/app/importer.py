from __future__ import annotations

from datetime import date, datetime, time
from pathlib import Path
import re

from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.models import EnergyConsumption

HEADER_PATTERN = re.compile(r"^\s*(\d+)\s+(.+?)\s*\(kWh\)\s*$")


def _parse_device_header(raw_header: str) -> tuple[str, str, str, str]:
    header = (raw_header or "").strip()
    match = HEADER_PATTERN.match(header)
    id_in_group = match.group(1) if match else "0"
    payload = match.group(2) if match else header
    parts = payload.split()

    # Надежный парсинг под формат "QF 1,26 ЗУ PzS 12V 1".
    if "ЗУ" in parts:
        zu_idx = parts.index("ЗУ")
        class_tokens = parts[:zu_idx]
        name_tokens = parts[zu_idx + 1 :]
        device_class = " ".join(class_tokens) if class_tokens else "UNKNOWN"
        device_type = "ЗУ"
        device_name = " ".join(name_tokens) if name_tokens else "UNKNOWN"
        return device_class, device_type, device_name, id_in_group

    if len(parts) >= 2:
        return " ".join(parts[:-1]), parts[-1], "UNKNOWN", id_in_group
    return "UNKNOWN", "UNKNOWN", payload, id_in_group


def _to_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        for fmt in ("%d.%m.%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                continue
    return None


def _to_time(value: object) -> time | None:
    if isinstance(value, datetime):
        return value.time().replace(second=0, microsecond=0)
    if isinstance(value, time):
        return value.replace(second=0, microsecond=0)
    if isinstance(value, str):
        for fmt in ("%H:%M:%S", "%H:%M"):
            try:
                return datetime.strptime(value.strip(), fmt).time()
            except ValueError:
                continue
    return None


def import_excel_if_empty(db: Session, excel_path: str) -> int:
    total_rows = db.scalar(select(func.count()).select_from(EnergyConsumption)) or 0
    if total_rows > 0:
        return 0

    path = Path(excel_path)
    if not path.exists():
        return 0

    wb = load_workbook(str(path), data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Во второй строке: дата, время, далее колонки оборудования.
    device_columns: list[tuple[int, tuple[str, str, str, str]]] = []
    for col_idx in range(3, ws.max_column + 1):
        raw_header = ws.cell(row=2, column=col_idx).value
        if raw_header is None:
            continue
        device_columns.append((col_idx, _parse_device_header(str(raw_header))))

    imported = 0
    batch: list[EnergyConsumption] = []
    for row_idx in range(3, ws.max_row + 1):
        row_date = _to_date(ws.cell(row=row_idx, column=1).value)
        row_time = _to_time(ws.cell(row=row_idx, column=2).value)
        if row_date is None or row_time is None:
            continue

        save_time = datetime.combine(row_date, row_time)
        for col_idx, (device_class, device_type, device_name, id_in_group) in device_columns:
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            try:
                energy_kwh = float(value)
            except (TypeError, ValueError):
                continue

            batch.append(
                EnergyConsumption(
                    device_class=device_class,
                    device_type=device_type,
                    device_name=device_name,
                    id_in_group=id_in_group,
                    save_time=save_time,
                    energy_kwh=energy_kwh,
                )
            )
            imported += 1

    if batch:
        db.add_all(batch)
        db.commit()

    return imported


def get_excel_time_bounds(excel_path: str) -> tuple[datetime | None, datetime | None]:
    path = Path(excel_path)
    if not path.exists():
        return None, None

    wb = load_workbook(str(path), data_only=True)
    ws = wb[wb.sheetnames[0]]

    min_time: datetime | None = None
    max_time: datetime | None = None
    for row_idx in range(3, ws.max_row + 1):
        row_date = _to_date(ws.cell(row=row_idx, column=1).value)
        row_time = _to_time(ws.cell(row=row_idx, column=2).value)
        if row_date is None or row_time is None:
            continue
        current = datetime.combine(row_date, row_time)
        if min_time is None or current < min_time:
            min_time = current
        if max_time is None or current > max_time:
            max_time = current
    return min_time, max_time
