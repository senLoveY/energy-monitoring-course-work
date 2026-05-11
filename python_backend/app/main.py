from __future__ import annotations

import asyncio
from collections import defaultdict
from datetime import date, datetime, timedelta, timezone
from io import BytesIO
from pathlib import Path
import json
import re
import smtplib
import threading
from email.message import EmailMessage

from fastapi import Depends, FastAPI, Form, Query, Request
from fastapi.responses import HTMLResponse, JSONResponse, PlainTextResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from fpdf import FPDF
import matplotlib
from matplotlib import font_manager
from matplotlib import pyplot as plt
from openpyxl import Workbook
from PIL import Image
from pydantic import BaseModel
from sqlalchemy import and_, asc, desc, func, select
from sqlalchemy.orm import Session
from starlette.middleware.sessions import SessionMiddleware

from app.auth import (
    consume_admin_invite,
    ensure_auth_schema,
    ensure_default_admin,
    generate_admin_invite_code,
    get_available_admin_invite,
    hash_password,
    verify_password,
)
from app.config import (
    TIMEZONE_OFFSET_MAX,
    TIMEZONE_OFFSET_MIN,
    get_timezone_offset_hours,
    set_timezone_offset_hours,
    settings,
)
from app.database import Base, engine
from app.database import SessionLocal, get_db
from app.importer import get_excel_time_bounds, import_excel_if_empty
from app.models import (
    AdminInviteCode,
    AppSetting,
    AppUser,
    AuditLog,
    EmailReportLog,
    EnergyConsumption,
    IncidentLog,
    ReportDispatchQueue,
    ReportSchedule,
)
from app.seed import seed_energy_data_if_empty
from app.services import EnergyConsumptionService, ReportDataNotFound
from app.timeutils import now_local_naive, today_local

matplotlib.use("Agg")

app = FastAPI(title="Energy Monitoring Python Backend")
app.add_middleware(SessionMiddleware, secret_key=settings.session_secret, same_site="lax")

TEMPLATES_DIR = Path(__file__).resolve().parent / "templates"
STATIC_DIR = Path(__file__).resolve().parent / "static"

templates = Jinja2Templates(directory=str(TEMPLATES_DIR))
app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")


def _safe_pdf_text(text: str) -> str:
    return str(text).encode("latin-1", "replace").decode("latin-1")


def _setup_pdf_font(pdf: FPDF) -> str:
    try:
        regular_path = font_manager.findfont("DejaVu Sans", fallback_to_default=True)
        bold_path = font_manager.findfont("DejaVu Sans:bold", fallback_to_default=True)
        pdf.add_font("DejaVu", style="", fname=regular_path)
        pdf.add_font("DejaVu", style="B", fname=bold_path)
        return "DejaVu"
    except Exception:
        return "Helvetica"


def _build_xlsx_export(payload: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Activity report"

    ws.append(["Report date", payload["report_date"]])
    ws.append(["Group", payload["selected_group"]])
    ws.append(["Total consumption (kWh)", payload["total_consumption"]])
    ws.append(["Peak (kWh)", payload["peak_consumption"]["value"]])
    ws.append(["Peak time", payload["peak_consumption"]["time"]])
    ws.append(["Low (kWh)", payload["low_consumption"]["value"]])
    ws.append(["Low time", payload["low_consumption"]["time"]])
    ws.append([])

    headers = ["Time"] + [device.name for device in payload["devices"]]
    ws.append(headers)
    for record in payload["consumption_records"]:
        ws.append([record["time"], *[item["value"] for item in record["consumptions"]]])

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream.getvalue()


def _build_plot_png(
    labels: list[str],
    values: list[float],
    title: str,
    kind: str = "line",
    width: float = 8.5,
    height: float = 3.2,
) -> bytes:
    fig, ax = plt.subplots(figsize=(width, height), dpi=130)
    if kind == "bar":
        ax.bar(labels, values, color="#3d63dd", alpha=0.8)
    elif kind == "pie":
        filtered = [(l, v) for l, v in zip(labels, values) if v > 0]
        if filtered:
            p_labels, p_values = zip(*filtered)
            ax.pie(p_values, labels=p_labels, autopct="%1.1f%%", startangle=80, textprops={"fontsize": 8})
            ax.axis("equal")
        else:
            ax.text(0.5, 0.5, "Нет данных", ha="center", va="center")
    else:
        ax.plot(labels, values, color="#3d63dd", linewidth=2.0)
        ax.fill_between(labels, values, color="#3d63dd", alpha=0.2)

    ax.set_title(title, fontsize=10, pad=10)
    if kind != "pie":
        ax.grid(True, axis="y", alpha=0.22)
        if labels:
            # Разрежаем подписи по X, чтобы время не накладывалось в PDF.
            max_ticks = 8
            step = max(1, len(labels) // max_ticks)
            tick_positions = list(range(0, len(labels), step))
            if tick_positions[-1] != len(labels) - 1:
                tick_positions.append(len(labels) - 1)
            tick_labels = [labels[i] for i in tick_positions]
            ax.set_xticks(tick_positions)
            ax.set_xticklabels(tick_labels)
        ax.tick_params(axis="x", labelsize=7, rotation=0)
        ax.tick_params(axis="y", labelsize=7)
    fig.tight_layout()

    stream = BytesIO()
    fig.savefig(stream, format="png")
    plt.close(fig)
    stream.seek(0)
    return stream.getvalue()


def _extract_report_series(payload: dict) -> dict:
    labels: list[str] = []
    total_by_time: list[float] = []
    active_by_time: list[int] = []
    peak_name_by_time: list[str] = []
    device_totals: dict[str, float] = defaultdict(float)

    for record in payload["consumption_records"]:
        labels.append(record["time"])
        consumptions = record["consumptions"]
        total = round(sum(float(item["value"] or 0.0) for item in consumptions), 3)
        total_by_time.append(total)
        active_by_time.append(sum(1 for item in consumptions if float(item["value"] or 0.0) > 0.0))

        peak_item = max(consumptions, key=lambda x: float(x["value"] or 0.0), default={"device_name": "-", "value": 0.0})
        peak_name_by_time.append(str(peak_item.get("device_name") or "-"))

        for item in consumptions:
            device_totals[str(item.get("device_name") or "-")] += float(item.get("value") or 0.0)

    return {
        "labels": labels,
        "total_by_time": total_by_time,
        "active_by_time": active_by_time,
        "peak_name_by_time": peak_name_by_time,
        "device_totals": device_totals,
    }


def _draw_image_fit(pdf: FPDF, img: Image.Image, x: float, y: float, box_w: float, box_h: float) -> None:
    img_w = float(img.width or 1)
    img_h = float(img.height or 1)
    scale = min(box_w / img_w, box_h / img_h)
    draw_w = img_w * scale
    draw_h = img_h * scale
    dx = x + (box_w - draw_w) / 2
    dy = y + (box_h - draw_h) / 2
    pdf.image(img, x=dx, y=dy, w=draw_w, h=draw_h)


EMAIL_PATTERN = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
TIMEZONE_SETTING_KEY = "timezone_offset_hours"
ROLE_LEVELS = {
    "observer": 10,
    "operator": 20,
    "user": 20,
    "engineer": 30,
    "admin": 40,
}
ROLE_CHOICES = ["observer", "operator", "engineer", "admin"]
ROLE_LABELS = {
    "observer": "Наблюдатель",
    "operator": "Оператор",
    "engineer": "Инженер",
    "admin": "Администратор",
    "user": "Оператор",
}
INCIDENT_TYPES = ["line_start", "maintenance", "failure", "manual_note"]
INCIDENT_SEVERITIES = ["low", "medium", "high", "critical"]
INCIDENT_TYPE_LABELS = {
    "line_start": "Пуск линии",
    "maintenance": "Ремонт",
    "failure": "Авария",
    "manual_note": "Ручная заметка",
}
INCIDENT_SEVERITY_LABELS = {
    "low": "Низкая",
    "medium": "Средняя",
    "high": "Высокая",
    "critical": "Критическая",
}


def _parse_timezone_offset(raw_value: str | int) -> int:
    try:
        offset = int(raw_value)
    except (TypeError, ValueError):
        raise ValueError("Некорректное смещение часового пояса")
    if offset < TIMEZONE_OFFSET_MIN or offset > TIMEZONE_OFFSET_MAX:
        raise ValueError(
            f"Смещение должно быть в диапазоне от {TIMEZONE_OFFSET_MIN} до {TIMEZONE_OFFSET_MAX}"
        )
    return offset


def _load_timezone_offset_setting(db: Session) -> int:
    setting = db.get(AppSetting, TIMEZONE_SETTING_KEY)
    if setting is None:
        offset = _parse_timezone_offset(get_timezone_offset_hours())
        db.add(AppSetting(key=TIMEZONE_SETTING_KEY, value=str(offset)))
        set_timezone_offset_hours(offset)
        return offset

    offset = _parse_timezone_offset(setting.value)
    set_timezone_offset_hours(offset)
    return offset


def _parse_measurement_save_time(raw_save_time: str | None) -> datetime:
    if not raw_save_time:
        return now_local_naive()
    try:
        parsed = datetime.fromisoformat(raw_save_time.replace("Z", "+00:00"))
    except ValueError:
        return now_local_naive()

    if parsed.tzinfo is None:
        return parsed
    local_tz = timezone(timedelta(hours=get_timezone_offset_hours()))
    return parsed.astimezone(local_tz).replace(tzinfo=None)


def _role_level(role: str | None) -> int:
    return ROLE_LEVELS.get((role or "").strip().lower(), 0)


def _role_label(role: str | None) -> str:
    return ROLE_LABELS.get((role or "").strip().lower(), role or "-")


def _has_min_role(user: AppUser, min_role: str) -> bool:
    return _role_level(user.role) >= _role_level(min_role)


def _public_user(user: AppUser | None) -> dict | None:
    if user is None:
        return None
    return {
        "id": int(user.id),
        "username": user.username,
        "role": user.role,
        "is_blocked": bool(user.is_blocked),
        "blocked_at": user.blocked_at.isoformat() if user.blocked_at else None,
        "created_at": user.created_at.isoformat() if user.created_at else None,
    }


def _render_react_page(
    request: Request,
    *,
    page: str,
    title: str,
    props: dict,
    style_paths: list[str],
    status_code: int = 200,
):
    payload_json = json.dumps({"page": page, "props": props}, ensure_ascii=False)
    return templates.TemplateResponse(
        request=request,
        name="react_app.html",
        context={"title": title, "style_paths": style_paths, "payload_json": payload_json},
        status_code=status_code,
    )


def _serialize_groups(groups: list) -> list[dict]:
    return [{"id": group.id, "label": group.label} for group in groups]


def _group_key_from_device_name(device_name: str | None) -> str:
    if not device_name:
        return "UNKNOWN"
    return device_name.split()[0].upper()


def _build_devices_by_group(db: Session, available_groups: list) -> dict[str, list[str]]:
    stmt = select(EnergyConsumption.device_name).distinct().order_by(asc(EnergyConsumption.device_name))
    names = [row[0] for row in db.execute(stmt).all() if row[0]]
    allowed_group_ids = {group.id for group in available_groups}
    devices_by_group: dict[str, list[str]] = {group.id: [] for group in available_groups}
    for name in names:
        group_id = _group_key_from_device_name(name)
        if group_id in allowed_group_ids:
            devices_by_group.setdefault(group_id, []).append(name)
    for group_id in list(devices_by_group.keys()):
        devices_by_group[group_id] = sorted(set(devices_by_group[group_id]))
    return devices_by_group


def _serialize_activity_payload(payload: dict) -> dict:
    return {
        "report_date": payload.get("report_date"),
        "selected_group": payload.get("selected_group", "all"),
        "available_groups": _serialize_groups(payload.get("available_groups", [])),
        "devices": [{"id": device.id, "name": device.name} for device in payload.get("devices", [])],
        "consumption_records_last_hour": payload.get("consumption_records_last_hour", []),
        "total_consumption": payload.get("total_consumption", 0),
        "peak_consumption": payload.get("peak_consumption", {"value": 0, "time": "-"}),
        "low_consumption": payload.get("low_consumption", {"value": 0, "time": "-"}),
        "trend": payload.get("trend", 0),
        "chart_labels": json.loads(payload.get("chart_labels_json", "[]")),
        "chart_datasets": json.loads(payload.get("chart_datasets_json", "[]")),
        "insights": payload.get("insights", {"recent_samples": 0, "group_stats": []}),
        "error_message": payload.get("error_message", ""),
        "timezone_offset_hours": payload.get("timezone_offset_hours", get_timezone_offset_hours()),
        "current_user": payload.get("current_user"),
        "current_user_role_label": payload.get("current_user_role_label", ""),
    }


def _serialize_incident(incident: IncidentLog) -> dict:
    return {
        "id": int(incident.id),
        "occurred_at": incident.occurred_at.strftime("%Y-%m-%d %H:%M"),
        "incident_type": incident.incident_type,
        "severity": incident.severity,
        "group_id": incident.group_id,
        "device_name": incident.device_name,
        "title": incident.title,
        "description": incident.description,
    }


def _serialize_admin_payload(
    *,
    users: list[AppUser],
    invite_codes: list[AdminInviteCode],
    schedules: list[ReportSchedule],
    email_logs: list[EmailReportLog],
    queue_items: list[ReportDispatchQueue],
    audit_logs: list[AuditLog],
    available_groups: list,
    selected_log_status: str,
    generated_code: str,
    schedule_message: str,
    settings_message: str,
    timezone_offset_hours: int,
    current_user: AppUser,
) -> dict:
    user_lookup = {int(user.id): user.username for user in users}

    def _parse_audit_details(raw_details: str | None) -> dict:
        if not raw_details:
            return {}
        try:
            decoded = json.loads(raw_details)
        except json.JSONDecodeError:
            return {"raw": raw_details}
        if isinstance(decoded, dict):
            return decoded
        return {"value": decoded}

    return {
        "users": [
            {
                "id": int(user.id),
                "username": user.username,
                "role": user.role,
                "is_blocked": bool(user.is_blocked),
                "created_at": user.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            }
            for user in users
        ],
        "invite_codes": [
            {
                "id": int(code.id),
                "code": code.code,
                "used_at": code.used_at.isoformat() if code.used_at else None,
            }
            for code in invite_codes
        ],
        "schedules": [
            {
                "id": int(item.id),
                "recipient_email": item.recipient_email,
                "group_id": item.group_id,
                "export_format": item.export_format,
                "template_mode": item.template_mode,
                "send_time": item.send_time,
                "is_enabled": bool(item.is_enabled),
            }
            for item in schedules
        ],
        "email_logs": [
            {
                "id": int(log.id),
                "created_at": log.created_at.strftime("%Y-%m-%d %H:%M:%S"),
                "recipient_email": log.recipient_email,
                "export_format": log.export_format,
                "status": log.status,
            }
            for log in email_logs
        ],
        "queue_items": [
            {
                "id": int(item.id),
                "recipient_email": item.recipient_email,
                "export_format": item.export_format,
                "report_date": item.report_date.isoformat(),
                "status": item.status,
            }
            for item in queue_items
        ],
        "audit_logs": [
            {
                "id": int(item.id),
                "created_at": item.created_at.strftime("%Y-%m-%d %H:%M:%S"),
                "actor_user_id": item.actor_user_id,
                "actor_username": user_lookup.get(int(item.actor_user_id)) if item.actor_user_id is not None else None,
                "action": item.action,
                "entity_type": item.entity_type,
                "entity_id": item.entity_id,
                "details": _parse_audit_details(item.details_json),
            }
            for item in audit_logs
        ],
        "available_groups": _serialize_groups(available_groups),
        "selected_log_status": selected_log_status,
        "generated_code": generated_code,
        "schedule_message": schedule_message,
        "settings_message": settings_message,
        "timezone_offset_hours": timezone_offset_hours,
        "role_choices": ROLE_CHOICES,
        "role_labels": ROLE_LABELS,
        "current_user_role_label": _role_label(current_user.role),
        "current_user": _public_user(current_user),
    }


def _load_incidents_for_day(db: Session, day: date, group: str, limit: int = 30) -> list[IncidentLog]:
    start = datetime.combine(day, datetime.min.time())
    end = start + timedelta(days=1)
    stmt = select(IncidentLog).where(IncidentLog.occurred_at >= start, IncidentLog.occurred_at < end)
    if group != "all":
        stmt = stmt.where(IncidentLog.group_id.in_([group, "all"]))
    stmt = stmt.order_by(desc(IncidentLog.occurred_at)).limit(limit)
    return list(db.scalars(stmt).all())


def _build_realtime_snapshot(db: Session, group: str = "all") -> dict:
    last_measurement_stmt = select(EnergyConsumption).order_by(desc(EnergyConsumption.save_time)).limit(1)
    if group != "all":
        last_measurement_stmt = (
            select(EnergyConsumption)
            .where(EnergyConsumption.device_name.ilike(f"{group}%"))
            .order_by(desc(EnergyConsumption.save_time))
            .limit(1)
        )
    last_measurement = db.scalar(last_measurement_stmt)
    now = now_local_naive()
    recent_since = now - timedelta(minutes=5)
    active_since = now - timedelta(minutes=1)
    total_stmt = select(func.coalesce(func.sum(EnergyConsumption.energy_kwh), 0.0)).where(
        EnergyConsumption.save_time >= recent_since
    )
    active_stmt = select(func.count(EnergyConsumption.id)).where(EnergyConsumption.save_time >= active_since)
    if group != "all":
        total_stmt = total_stmt.where(EnergyConsumption.device_name.ilike(f"{group}%"))
        active_stmt = active_stmt.where(EnergyConsumption.device_name.ilike(f"{group}%"))

    total_last_5m = float(db.scalar(total_stmt) or 0.0)
    active_samples_1m = int(db.scalar(active_stmt) or 0)
    return {
        "last_seen": last_measurement.save_time.strftime("%Y-%m-%d %H:%M:%S") if last_measurement else None,
        "group": group,
        "total_last_5m": round(total_last_5m, 3),
        "active_samples_1m": active_samples_1m,
        "server_time": now.strftime("%Y-%m-%d %H:%M:%S"),
    }


def _smtp_is_configured() -> bool:
    return bool(settings.mail_host and settings.mail_password and (settings.mail_username or settings.mail_from))


def _send_report_email(
    recipient_email: str,
    report_date: date,
    group: str,
    export_format: str,
    content: bytes,
) -> None:
    attachment_name = f"activity-{report_date.isoformat()}-{group or 'all'}.{export_format}"
    content_type = (
        "application/pdf"
        if export_format == "pdf"
        else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    msg = EmailMessage()
    msg["From"] = settings.mail_from
    msg["To"] = recipient_email
    msg["Subject"] = f"Отчет по энергопотреблению за {report_date.isoformat()}"
    msg.set_content(
        "\n".join(
            [
                "Здравствуйте!",
                "",
                "Во вложении находится сформированный отчет по энергопотреблению.",
                f"Дата отчета: {report_date.isoformat()}",
                f"Группа устройств: {group or 'all'}",
            ]
        )
    )
    maintype, subtype = content_type.split("/", maxsplit=1)
    msg.add_attachment(content, maintype=maintype, subtype=subtype, filename=attachment_name)

    smtp_login = settings.mail_username or settings.mail_from.split("@", maxsplit=1)[0]
    encryption = (settings.mail_encryption or "").lower()
    use_ssl = encryption == "ssl" or settings.mail_port == 465

    if use_ssl:
        with smtplib.SMTP_SSL(settings.mail_host, settings.mail_port, timeout=20) as server:
            server.ehlo()
            server.login(smtp_login, settings.mail_password)
            server.send_message(msg)
        return

    with smtplib.SMTP(settings.mail_host, settings.mail_port, timeout=20) as server:
        server.ehlo()
        if encryption in {"tls", "starttls"} or settings.mail_port in {587, 25}:
            server.starttls()
            server.ehlo()
        server.login(smtp_login, settings.mail_password)
        server.send_message(msg)


def _build_export_bytes(payload: dict, export_format: str, template_mode: str = "detailed") -> bytes:
    if export_format == "xlsx":
        return _build_xlsx_export(payload)
    return _build_pdf_export(payload, template_mode=template_mode)


def _log_audit(
    db: Session,
    *,
    actor_user_id: int | None,
    action: str,
    entity_type: str,
    entity_id: str | None = None,
    details: dict | None = None,
) -> None:
    db.add(
        AuditLog(
            actor_user_id=actor_user_id,
            action=action,
            entity_type=entity_type,
            entity_id=entity_id,
            details_json=json.dumps(details or {}, ensure_ascii=False),
        )
    )


def _queue_report_send(
    db: Session,
    *,
    recipient_email: str,
    report_date: date,
    group: str,
    export_format: str,
    template_mode: str,
    requested_by_user_id: int | None,
    schedule_id: int | None = None,
) -> ReportDispatchQueue:
    queue_item = ReportDispatchQueue(
        recipient_email=recipient_email,
        report_date=report_date,
        group_id=group,
        export_format=export_format,
        template_mode=template_mode,
        status="pending",
        attempts=0,
        max_attempts=3,
        next_attempt_at=now_local_naive(),
        requested_by_user_id=requested_by_user_id,
        schedule_id=schedule_id,
    )
    db.add(queue_item)
    db.flush()
    db.add(
        EmailReportLog(
            queue_item_id=queue_item.id,
            requested_by_user_id=requested_by_user_id,
            recipient_email=recipient_email,
            report_date=report_date,
            group_id=group,
            export_format=export_format,
            template_mode=template_mode,
            status="queued",
        )
    )
    return queue_item


def _build_pdf_export(payload: dict, template_mode: str = "detailed") -> bytes:
    series = _extract_report_series(payload)
    top_devices = sorted(series["device_totals"].items(), key=lambda x: x[1], reverse=True)[:8]
    pie_labels = [name if len(name) <= 14 else f"{name[:13]}…" for name, _ in top_devices]
    pie_values = [round(value, 3) for _, value in top_devices]
    bar_labels = pie_labels[:6]
    bar_values = pie_values[:6]

    line_chart = _build_plot_png(
        labels=series["labels"],
        values=series["total_by_time"],
        title="Суммарная нагрузка по времени",
        kind="line",
        width=7.2,
        height=2.9,
    )
    pie_chart = _build_plot_png(
        labels=pie_labels or ["Нет данных"],
        values=pie_values or [1.0],
        title="Распределение по устройствам",
        kind="pie",
        width=3.6,
        height=2.9,
    )
    bar_chart = _build_plot_png(
        labels=bar_labels,
        values=bar_values,
        title="Топ устройств (кВт*ч)",
        kind="bar",
        width=7.8,
        height=2.5,
    )

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.add_page()
    font_family = _setup_pdf_font(pdf)
    pdf.set_fill_color(38, 56, 88)
    pdf.rect(0, 0, 210, 26, style="F")
    pdf.set_text_color(255, 255, 255)
    pdf.set_font(font_family, "B", 16)
    pdf.set_xy(10, 8)
    pdf.cell(0, 8, "Отчет по энергопотреблению", ln=True)
    pdf.set_font(font_family, size=9)
    pdf.set_xy(10, 17)
    pdf.cell(0, 5, f"Дата: {payload['report_date']}  |  Группа: {payload['selected_group']}", ln=True)
    pdf.set_text_color(30, 39, 58)
    pdf.ln(8)

    card_y = pdf.get_y()
    cards = [
        ("Суммарно (кВт*ч)", str(payload["total_consumption"])),
        ("Пиковая нагрузка", f"{payload['peak_consumption']['value']} в {payload['peak_consumption']['time']}"),
        ("Минимальная нагрузка", f"{payload['low_consumption']['value']} в {payload['low_consumption']['time']}"),
    ]
    card_w = 61
    gap = 3
    for idx, (title, value) in enumerate(cards):
        x = 10 + idx * (card_w + gap)
        pdf.set_xy(x, card_y)
        pdf.set_fill_color(244, 247, 252)
        pdf.set_draw_color(209, 217, 232)
        pdf.rect(x, card_y, card_w, 18, style="DF")
        pdf.set_xy(x + 2, card_y + 2)
        pdf.set_font(font_family, size=8)
        pdf.cell(card_w - 4, 4, title, ln=True)
        pdf.set_xy(x + 2, card_y + 8)
        pdf.set_font(font_family, "B", 10)
        pdf.cell(card_w - 4, 6, value, ln=True)
    pdf.ln(22)

    if template_mode == "brief":
        chart_top = pdf.get_y()
        line_img = Image.open(BytesIO(line_chart))
        pdf.set_draw_color(209, 217, 232)
        pdf.set_fill_color(255, 255, 255)
        pdf.rect(10, chart_top, 190, 72, style="DF")
        _draw_image_fit(pdf, line_img, x=10, y=chart_top, box_w=190, box_h=72)
        pdf.set_y(chart_top + 76)
        pdf.set_font(font_family, "B", 11)
        pdf.cell(0, 8, "Ключевые показатели по времени", ln=True)
        pdf.set_font(font_family, size=9)
        tail_labels = series["labels"][-12:]
        tail_values = series["total_by_time"][-12:]
        for label, value in zip(tail_labels, tail_values):
            pdf.cell(0, 6, f"{label}: {value:.2f} кВт*ч", ln=True)
        return bytes(pdf.output(dest="S"))

    chart_top = pdf.get_y()
    line_img = Image.open(BytesIO(line_chart))
    pie_img = Image.open(BytesIO(pie_chart))
    bar_img = Image.open(BytesIO(bar_chart))

    # Рамки под графики фиксированные, а сами графики вписываются пропорционально.
    pdf.set_draw_color(209, 217, 232)
    pdf.set_fill_color(255, 255, 255)
    pdf.rect(10, chart_top, 130, 56, style="DF")
    pdf.rect(145, chart_top, 55, 56, style="DF")
    _draw_image_fit(pdf, line_img, x=10, y=chart_top, box_w=130, box_h=56)
    _draw_image_fit(pdf, pie_img, x=145, y=chart_top, box_w=55, box_h=56)

    second_row_y = chart_top + 60
    pdf.rect(10, second_row_y, 190, 45, style="DF")
    _draw_image_fit(pdf, bar_img, x=10, y=second_row_y, box_w=190, box_h=45)
    pdf.set_y(second_row_y)
    pdf.ln(48)

    pdf.set_font(font_family, "B", 11)
    pdf.cell(0, 8, "Таблица потребления (последние записи)", ln=True)
    pdf.set_font(font_family, "B", 9)
    pdf.set_fill_color(235, 240, 248)
    pdf.set_draw_color(209, 217, 232)
    row_h = 7
    w_time, w_total, w_active, w_load = 28, 34, 28, 100
    pdf.cell(w_time, row_h, "Время", border=1, fill=True)
    pdf.cell(w_total, row_h, "Сумма", border=1, fill=True)
    pdf.cell(w_active, row_h, "Актив.", border=1, fill=True)
    pdf.cell(w_load, row_h, "Уровень", border=1, fill=True, ln=True)

    max_rows = 95
    total_max = max(series["total_by_time"], default=1.0) or 1.0
    pdf.set_font(font_family, size=8)
    for idx, label in enumerate(series["labels"][:max_rows]):
        if pdf.get_y() > 276:
            pdf.add_page()
            pdf.set_font(font_family, "B", 9)
            pdf.cell(w_time, row_h, "Время", border=1, fill=True)
            pdf.cell(w_total, row_h, "Сумма", border=1, fill=True)
            pdf.cell(w_active, row_h, "Актив.", border=1, fill=True)
            pdf.cell(w_load, row_h, "Уровень", border=1, fill=True, ln=True)
            pdf.set_font(font_family, size=8)

        total = series["total_by_time"][idx]
        active = series["active_by_time"][idx]

        x = pdf.get_x()
        y = pdf.get_y()
        pdf.cell(w_time, row_h, label, border=1)
        pdf.cell(w_total, row_h, f"{total:.2f}", border=1)
        pdf.cell(w_active, row_h, str(active), border=1)
        pdf.cell(w_load, row_h, "", border=1, ln=True)

        bar_ratio = max(0.0, min(1.0, total / total_max))
        bar_w = (w_load - 4) * bar_ratio
        pdf.set_fill_color(61, 99, 221)
        pdf.rect(x + w_time + w_total + w_active + 2, y + 2, bar_w, row_h - 4, style="F")

    if len(series["labels"]) > max_rows:
        pdf.ln(2)
        pdf.set_font(font_family, size=8)
        pdf.cell(0, 6, "Таблица сокращена для печатного формата.", ln=True)

    return bytes(pdf.output(dest="S"))


RETRY_DELAYS_SECONDS = [60, 300, 900]


def _process_schedules(db: Session) -> int:
    now = now_local_naive()
    current_hhmm = now.strftime("%H:%M")
    today = now.date()
    schedules = list(
        db.scalars(
            select(ReportSchedule).where(
                ReportSchedule.is_enabled.is_(True),
                ReportSchedule.send_time == current_hhmm,
            )
        ).all()
    )
    created = 0
    for schedule in schedules:
        if schedule.last_run_on == today:
            continue
        formats = ["pdf", "xlsx"] if schedule.export_format == "both" else [schedule.export_format]
        for fmt in formats:
            _queue_report_send(
                db,
                recipient_email=schedule.recipient_email,
                report_date=today,
                group=schedule.group_id,
                export_format=fmt,
                template_mode=schedule.template_mode,
                requested_by_user_id=schedule.created_by_user_id,
                schedule_id=schedule.id,
            )
            created += 1
        schedule.last_run_on = today
        _log_audit(
            db,
            actor_user_id=schedule.created_by_user_id,
            action="schedule_triggered",
            entity_type="report_schedule",
            entity_id=str(schedule.id),
            details={"queued_jobs": created, "run_on": str(today)},
        )
    return created


def _process_queue(db: Session) -> int:
    now = now_local_naive()
    items = list(
        db.scalars(
            select(ReportDispatchQueue)
            .where(
                and_(
                    ReportDispatchQueue.status.in_(["pending", "retry"]),
                    ReportDispatchQueue.next_attempt_at <= now,
                )
            )
            .order_by(asc(ReportDispatchQueue.created_at))
            .limit(20)
        ).all()
    )
    processed = 0
    for item in items:
        item.status = "sending"
        db.flush()
        try:
            service = EnergyConsumptionService(db)
            report_payload = service.build_activity_report(item.report_date, group=item.group_id)
            content = _build_export_bytes(report_payload, item.export_format, template_mode=item.template_mode)
            _send_report_email(
                recipient_email=item.recipient_email,
                report_date=item.report_date,
                group=item.group_id,
                export_format=item.export_format,
                content=content,
            )
            item.status = "sent"
            item.sent_at = now_local_naive()
            item.last_error = None
            db.add(
                EmailReportLog(
                    queue_item_id=item.id,
                    requested_by_user_id=item.requested_by_user_id,
                    recipient_email=item.recipient_email,
                    report_date=item.report_date,
                    group_id=item.group_id,
                    export_format=item.export_format,
                    template_mode=item.template_mode,
                    status="sent",
                    sent_at=item.sent_at,
                )
            )
            _log_audit(
                db,
                actor_user_id=item.requested_by_user_id,
                action="report_sent",
                entity_type="report_dispatch_queue",
                entity_id=str(item.id),
                details={"email": item.recipient_email, "format": item.export_format},
            )
        except Exception as exc:
            item.attempts += 1
            item.last_error = str(exc)
            if item.attempts >= item.max_attempts:
                item.status = "failed"
                item.next_attempt_at = now_local_naive() + timedelta(days=3650)
            else:
                item.status = "retry"
                delay = RETRY_DELAYS_SECONDS[min(item.attempts - 1, len(RETRY_DELAYS_SECONDS) - 1)]
                item.next_attempt_at = now_local_naive() + timedelta(seconds=delay)
            db.add(
                EmailReportLog(
                    queue_item_id=item.id,
                    requested_by_user_id=item.requested_by_user_id,
                    recipient_email=item.recipient_email,
                    report_date=item.report_date,
                    group_id=item.group_id,
                    export_format=item.export_format,
                    template_mode=item.template_mode,
                    status=item.status,
                    error_message=item.last_error,
                )
            )
        processed += 1
    return processed


def _queue_worker_loop(stop_event: threading.Event) -> None:
    while not stop_event.is_set():
        try:
            with SessionLocal() as db:
                _process_schedules(db)
                _process_queue(db)
                db.commit()
        except Exception:
            pass
        stop_event.wait(20)


@app.on_event("startup")
def startup() -> None:
    Base.metadata.create_all(bind=engine)
    with SessionLocal() as db:
        app.state.excel_max_time = None
        app.state.timezone_offset_hours = _load_timezone_offset_setting(db)
        ensure_auth_schema(db)
        ensure_default_admin(db)
        if Path(settings.excel_data_path).exists():
            _, excel_max_time = get_excel_time_bounds(settings.excel_data_path)
            app.state.excel_max_time = excel_max_time
            imported = import_excel_if_empty(db, settings.excel_data_path)
            if imported == 0 and (db.query(EnergyConsumption).count() == 0):
                seed_energy_data_if_empty(db)
        else:
            if db.query(EnergyConsumption).count() == 0:
                seed_energy_data_if_empty(db)
    stop_event = threading.Event()
    worker = threading.Thread(target=_queue_worker_loop, args=(stop_event,), daemon=True, name="report-queue-worker")
    worker.start()
    app.state.queue_stop_event = stop_event
    app.state.queue_worker = worker


@app.on_event("shutdown")
def shutdown() -> None:
    stop_event = getattr(app.state, "queue_stop_event", None)
    worker = getattr(app.state, "queue_worker", None)
    if stop_event is not None:
        stop_event.set()
    if worker is not None and worker.is_alive():
        worker.join(timeout=2)


def get_current_user(request: Request, db: Session) -> AppUser | None:
    user_id = request.session.get("user_id")
    if user_id is None:
        return None
    user = db.get(AppUser, user_id)
    if user is None or user.is_blocked:
        return None
    return user


def require_user(request: Request, db: Session) -> AppUser | RedirectResponse:
    user = get_current_user(request, db)
    if user is None:
        request.session.clear()
        return RedirectResponse(url="/login", status_code=303)
    return user


def require_admin(request: Request, db: Session) -> AppUser | RedirectResponse:
    user = get_current_user(request, db)
    if user is None:
        request.session.clear()
        return RedirectResponse(url="/login", status_code=303)
    if user.role != "admin":
        return RedirectResponse(url="/activity", status_code=303)
    return user


def require_min_role(request: Request, db: Session, min_role: str) -> AppUser | RedirectResponse:
    user = get_current_user(request, db)
    if user is None:
        request.session.clear()
        return RedirectResponse(url="/login", status_code=303)
    if not _has_min_role(user, min_role):
        return RedirectResponse(url="/activity", status_code=303)
    return user


class MeasurementIn(BaseModel):
    device_class: str
    device_type: str
    device_name: str
    id_in_group: str
    save_time: str | None = None
    energy_kwh: float


class EmailReportIn(BaseModel):
    email: str
    report_date: str
    group: str = "all"
    export_format: str = "pdf"
    template_mode: str = "detailed"


class ReportScheduleIn(BaseModel):
    email: str
    group: str = "all"
    send_time: str = "08:00"
    export_format: str = "both"
    template_mode: str = "detailed"


@app.post("/ingest/measurement")
def ingest_measurement(payload: MeasurementIn, db: Session = Depends(get_db)) -> dict:
    actual_save_time = _parse_measurement_save_time(payload.save_time)

    db.add(
        EnergyConsumption(
            device_class=payload.device_class,
            device_type=payload.device_type,
            device_name=payload.device_name,
            id_in_group=payload.id_in_group,
            save_time=actual_save_time,
            energy_kwh=payload.energy_kwh,
        )
    )
    db.commit()
    return {"status": "accepted"}


@app.get("/", response_class=HTMLResponse)
def root(request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request, db)
    return _render_react_page(
        request,
        page="landing",
        title="Energy Monitoring",
        props={"current_user": _public_user(user), "current_user_role_label": _role_label(user.role) if user else ""},
        style_paths=["styles/landing.css"],
    )


@app.get("/login", response_class=HTMLResponse)
def login_page(request: Request, db: Session = Depends(get_db)):
    if get_current_user(request, db) is not None:
        return RedirectResponse(url="/activity", status_code=303)
    return _render_react_page(
        request,
        page="login",
        title="Вход",
        props={"error": ""},
        style_paths=["styles/auth.css"],
    )


@app.post("/login", response_class=HTMLResponse)
async def login_action(request: Request, db: Session = Depends(get_db)):
    form = await request.form()
    username = str(form.get("username", "")).strip()
    password = str(form.get("password", ""))

    stmt = select(AppUser).where(AppUser.username == username)
    user = db.scalar(stmt)
    if user is None or not verify_password(password, user.password_hash):
        return _render_react_page(
            request,
            page="login",
            title="Вход",
            props={"error": "Неверный логин или пароль"},
            style_paths=["styles/auth.css"],
            status_code=400,
        )
    if user.is_blocked:
        return _render_react_page(
            request,
            page="login",
            title="Вход",
            props={"error": "Пользователь заблокирован. Обратитесь к администратору."},
            style_paths=["styles/auth.css"],
            status_code=403,
        )

    request.session["user_id"] = user.id
    return RedirectResponse(url="/activity", status_code=303)


@app.get("/register", response_class=HTMLResponse)
def register_page(request: Request, db: Session = Depends(get_db)):
    if get_current_user(request, db) is not None:
        return RedirectResponse(url="/activity", status_code=303)
    return _render_react_page(
        request,
        page="register",
        title="Регистрация",
        props={"error": ""},
        style_paths=["styles/auth.css"],
    )


@app.post("/register", response_class=HTMLResponse)
async def register_action(request: Request, db: Session = Depends(get_db)):
    form = await request.form()
    username = str(form.get("username", "")).strip()
    password = str(form.get("password", ""))
    role = str(form.get("role", "observer")).strip().lower()
    admin_code = str(form.get("admin_code", "")).strip()

    if not username or len(password) < 6:
        return _render_react_page(
            request,
            page="register",
            title="Регистрация",
            props={"error": "Введите username и пароль минимум из 6 символов"},
            style_paths=["styles/auth.css"],
            status_code=400,
        )
    if role not in set(ROLE_CHOICES):
        role = "observer"
    invite: AdminInviteCode | None = None
    if role == "admin":
        if admin_code == settings.admin_invite_code:
            invite = None
        else:
            invite = get_available_admin_invite(db, admin_code)
            if invite is None:
                return _render_react_page(
                    request,
                    page="register",
                    title="Регистрация",
                    props={"error": "Неверный или уже использованный код администратора"},
                    style_paths=["styles/auth.css"],
                    status_code=400,
                )

    existing = db.scalar(select(AppUser).where(AppUser.username == username))
    if existing is not None:
        return _render_react_page(
            request,
            page="register",
            title="Регистрация",
            props={"error": "Пользователь с таким username уже существует"},
            style_paths=["styles/auth.css"],
            status_code=400,
        )

    user = AppUser(username=username, password_hash=hash_password(password), role=role)
    db.add(user)
    db.flush()
    if invite is not None:
        consume_admin_invite(db, invite, used_by_user_id=user.id)
    db.commit()
    db.refresh(user)
    request.session["user_id"] = user.id
    return RedirectResponse(url="/activity", status_code=303)


@app.get("/logout")
def logout(request: Request):
    request.session.clear()
    return RedirectResponse(url="/login", status_code=303)


@app.get("/health")
def healthcheck() -> dict[str, str]:
    return {"status": "ok"}


@app.get("/activity/stream")
async def activity_stream(request: Request, group: str = Query(default="all"), db: Session = Depends(get_db)):
    current_user = require_user(request, db)
    if isinstance(current_user, RedirectResponse):
        return PlainTextResponse("unauthorized", status_code=401)

    async def event_generator():
        last_payload = ""
        while True:
            if await request.is_disconnected():
                break
            with SessionLocal() as stream_db:
                snapshot = _build_realtime_snapshot(stream_db, group=group)
            payload = json.dumps(snapshot, ensure_ascii=False)
            if payload != last_payload:
                yield f"event: stats\ndata: {payload}\n\n"
                last_payload = payload
            else:
                yield "event: ping\ndata: {}\n\n"
            await asyncio.sleep(5)

    headers = {
        "Cache-Control": "no-cache",
        "Connection": "keep-alive",
        "X-Accel-Buffering": "no",
    }
    return StreamingResponse(event_generator(), media_type="text/event-stream", headers=headers)


@app.get("/activity", response_class=HTMLResponse)
def activity_report(
    request: Request,
    report_date: date | None = Query(default=None, alias="date"),
    group: str = Query(default="all"),
    db: Session = Depends(get_db),
):
    current_user = require_user(request, db)
    if isinstance(current_user, RedirectResponse):
        return current_user

    day = report_date or today_local()
    service = EnergyConsumptionService(db)
    available_groups = service.get_available_groups()
    insights = service.build_data_insights(getattr(request.app.state, "excel_max_time", None))
    try:
        payload = service.build_activity_report(day, group=group)
    except ReportDataNotFound as exc:
        return _render_react_page(
            request,
            page="activity",
            title="Отчеты о статусах активности",
            props=_serialize_activity_payload(
                {
                    "current_user": _public_user(current_user),
                    "current_user_role_label": _role_label(current_user.role),
                    "report_date": day.isoformat(),
                    "selected_group": group,
                    "available_groups": available_groups,
                    "insights": insights,
                    "error_message": str(exc),
                    "devices": [],
                    "consumption_records_last_hour": [],
                    "total_consumption": 0,
                    "peak_consumption": {"value": 0, "time": "-"},
                    "low_consumption": {"value": 0, "time": "-"},
                    "trend": 0,
                    "chart_labels_json": "[]",
                    "chart_datasets_json": "[]",
                    "timezone_offset_hours": get_timezone_offset_hours(),
                }
            ),
            style_paths=["styles/activity.css"],
        )

    payload["current_user"] = _public_user(current_user)
    payload["current_user_role_label"] = _role_label(current_user.role)
    payload["insights"] = insights
    payload["error_message"] = ""
    payload["timezone_offset_hours"] = get_timezone_offset_hours()
    return _render_react_page(
        request,
        page="activity",
        title="Отчеты о статусах активности",
        props=_serialize_activity_payload(payload),
        style_paths=["styles/activity.css"],
    )


@app.get("/incidents", response_class=HTMLResponse)
def incidents_page(
    request: Request,
    report_date: date | None = Query(default=None, alias="date"),
    group: str = Query(default="all"),
    db: Session = Depends(get_db),
):
    current_user = require_user(request, db)
    if isinstance(current_user, RedirectResponse):
        return current_user

    day = report_date or today_local()
    service = EnergyConsumptionService(db)
    available_groups = service.get_available_groups()
    devices_by_group = _build_devices_by_group(db, available_groups)
    incidents = _load_incidents_for_day(db, day=day, group=group, limit=100)
    return _render_react_page(
        request,
        page="incidents",
        title="Журнал инцидентов",
        props={
            "current_user": _public_user(current_user),
            "current_user_role_label": _role_label(current_user.role),
            "report_date": day.isoformat(),
            "selected_group": group,
            "available_groups": _serialize_groups(available_groups),
            "devices_by_group": devices_by_group,
            "timezone_offset_hours": get_timezone_offset_hours(),
            "incidents": [_serialize_incident(item) for item in incidents],
            "incident_types": INCIDENT_TYPES,
            "incident_severities": INCIDENT_SEVERITIES,
            "incident_type_labels": INCIDENT_TYPE_LABELS,
            "incident_severity_labels": INCIDENT_SEVERITY_LABELS,
            "can_manage_incidents": _has_min_role(current_user, "operator"),
            "incident_message": request.query_params.get("incident_message", ""),
        },
        style_paths=["styles/activity.css"],
    )


@app.get("/activity/export")
def export_activity_report(
    request: Request,
    report_date: date | None = Query(default=None, alias="date"),
    group: str = Query(default="all"),
    export_format: str = Query(default="xlsx", alias="format"),
    template_mode: str = Query(default="detailed", alias="template"),
    db: Session = Depends(get_db),
):
    current_user = require_user(request, db)
    if isinstance(current_user, RedirectResponse):
        return current_user

    day = report_date or today_local()
    service = EnergyConsumptionService(db)

    try:
        payload = service.build_activity_report(day, group=group)
    except ReportDataNotFound as exc:
        return PlainTextResponse(str(exc), status_code=404)

    safe_group = group if group else "all"
    if export_format == "xlsx":
        content = _build_export_bytes(payload, "xlsx", template_mode=template_mode)
        filename = f"activity-{day.isoformat()}-{safe_group}.xlsx"
        _log_audit(
            db,
            actor_user_id=current_user.id,
            action="report_exported",
            entity_type="activity_report",
            entity_id=day.isoformat(),
            details={"group": safe_group, "format": "xlsx", "template": template_mode},
        )
        db.commit()
        return StreamingResponse(
            BytesIO(content),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    if export_format == "pdf":
        content = _build_export_bytes(payload, "pdf", template_mode=template_mode)
        filename = f"activity-{day.isoformat()}-{safe_group}.pdf"
        _log_audit(
            db,
            actor_user_id=current_user.id,
            action="report_exported",
            entity_type="activity_report",
            entity_id=day.isoformat(),
            details={"group": safe_group, "format": "pdf", "template": template_mode},
        )
        db.commit()
        return StreamingResponse(
            BytesIO(content),
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    return PlainTextResponse("Unsupported export format", status_code=400)


@app.post("/activity/email-report")
def email_activity_report(payload_in: EmailReportIn, request: Request, db: Session = Depends(get_db)):
    current_user = require_min_role(request, db, "engineer")
    if isinstance(current_user, RedirectResponse):
        return JSONResponse({"error": "Недостаточно прав для отправки отчетов"}, status_code=403)

    if not EMAIL_PATTERN.match(payload_in.email.strip()):
        return JSONResponse({"error": "Некорректный email"}, status_code=400)
    if payload_in.export_format not in {"pdf", "xlsx"}:
        return JSONResponse({"error": "Неподдерживаемый формат"}, status_code=400)
    if payload_in.template_mode not in {"brief", "detailed"}:
        return JSONResponse({"error": "Неподдерживаемый шаблон"}, status_code=400)
    if not _smtp_is_configured():
        return JSONResponse({"error": "SMTP не настроен на сервере"}, status_code=500)

    try:
        day = date.fromisoformat(payload_in.report_date)
    except ValueError:
        return JSONResponse({"error": "Некорректная дата отчета"}, status_code=400)

    group = (payload_in.group or "all").strip() or "all"
    _queue_report_send(
        db,
        recipient_email=payload_in.email.strip(),
        report_date=day,
        group=group,
        export_format=payload_in.export_format,
        template_mode=payload_in.template_mode,
        requested_by_user_id=current_user.id,
    )
    _log_audit(
        db,
        actor_user_id=current_user.id,
        action="report_email_enqueued",
        entity_type="activity_report",
        entity_id=day.isoformat(),
        details={
            "group": group,
            "email": payload_in.email.strip(),
            "format": payload_in.export_format,
            "template": payload_in.template_mode,
        },
    )
    db.commit()

    return {"status": "ok", "message": "Письмо поставлено в очередь отправки"}


@app.post("/activity/schedules")
def create_report_schedule(payload_in: ReportScheduleIn, request: Request, db: Session = Depends(get_db)):
    current_user = require_min_role(request, db, "engineer")
    if isinstance(current_user, RedirectResponse):
        return JSONResponse({"error": "Недостаточно прав для создания расписаний"}, status_code=403)

    if not EMAIL_PATTERN.match(payload_in.email.strip()):
        return JSONResponse({"error": "Некорректный email"}, status_code=400)
    if not re.match(r"^\d{2}:\d{2}$", payload_in.send_time):
        return JSONResponse({"error": "Время должно быть в формате HH:MM"}, status_code=400)
    if payload_in.export_format not in {"pdf", "xlsx", "both"}:
        return JSONResponse({"error": "Неподдерживаемый формат расписания"}, status_code=400)
    if payload_in.template_mode not in {"brief", "detailed"}:
        return JSONResponse({"error": "Неподдерживаемый шаблон"}, status_code=400)

    schedule = ReportSchedule(
        recipient_email=payload_in.email.strip(),
        group_id=(payload_in.group or "all").strip() or "all",
        export_format=payload_in.export_format,
        template_mode=payload_in.template_mode,
        send_time=payload_in.send_time,
        is_enabled=True,
        created_by_user_id=current_user.id,
    )
    db.add(schedule)
    db.flush()
    _log_audit(
        db,
        actor_user_id=current_user.id,
        action="schedule_created",
        entity_type="report_schedule",
        entity_id=str(schedule.id),
        details={
            "email": schedule.recipient_email,
            "group": schedule.group_id,
            "format": schedule.export_format,
            "time": schedule.send_time,
            "template": schedule.template_mode,
        },
    )
    db.commit()
    return {"status": "ok", "message": "Ежедневная рассылка создана", "schedule_id": schedule.id}


@app.post("/activity/incidents")
@app.post("/incidents")
async def create_incident(request: Request, db: Session = Depends(get_db)):
    current_user = require_min_role(request, db, "operator")
    if isinstance(current_user, RedirectResponse):
        return current_user

    form = await request.form()
    report_date = str(form.get("date", "")).strip() or today_local().isoformat()
    page_group = str(form.get("group", "all")).strip() or "all"
    incident_group = str(form.get("incident_group", page_group)).strip() or page_group
    incident_device_name = str(form.get("incident_device_name", "__all__")).strip() or "__all__"
    incident_type = str(form.get("incident_type", "manual_note")).strip()
    severity = str(form.get("severity", "medium")).strip()
    occurred_at_raw = str(form.get("occurred_at", "")).strip()
    title = str(form.get("title", "")).strip()
    description = str(form.get("description", "")).strip()

    service = EnergyConsumptionService(db)
    available_groups = service.get_available_groups()
    available_group_ids = {group.id for group in available_groups}
    devices_by_group = _build_devices_by_group(db, available_groups)

    if incident_group not in available_group_ids and incident_group != "all":
        incident_group = page_group if (page_group in available_group_ids or page_group == "all") else "all"

    if not title:
        return RedirectResponse(
            url=f"/incidents?date={report_date}&group={page_group}&incident_message=Укажите+название+инцидента",
            status_code=303,
        )
    if incident_device_name != "__all__":
        if incident_group == "all":
            all_devices = {name for names in devices_by_group.values() for name in names}
            if incident_device_name not in all_devices:
                return RedirectResponse(
                    url=f"/incidents?date={report_date}&group={page_group}&incident_message=Выбранное+устройство+не+найдено",
                    status_code=303,
                )
        else:
            group_devices = set(devices_by_group.get(incident_group, []))
            if incident_device_name not in group_devices:
                return RedirectResponse(
                    url=f"/incidents?date={report_date}&group={page_group}&incident_message=Выбранное+устройство+не+принадлежит+выбранной+группе",
                    status_code=303,
                )

    if incident_type not in INCIDENT_TYPES:
        incident_type = "manual_note"
    if severity not in INCIDENT_SEVERITIES:
        severity = "medium"
    try:
        occurred_at = datetime.fromisoformat(occurred_at_raw) if occurred_at_raw else now_local_naive()
    except ValueError:
        occurred_at = now_local_naive()

    incident = IncidentLog(
        occurred_at=occurred_at,
        incident_type=incident_type,
        severity=severity,
        group_id=incident_group,
        device_name=None if incident_device_name == "__all__" else incident_device_name,
        title=title,
        description=description or None,
        created_by_user_id=current_user.id,
    )
    db.add(incident)
    db.flush()
    _log_audit(
        db,
        actor_user_id=current_user.id,
        action="incident_created",
        entity_type="incident_log",
        entity_id=str(incident.id),
        details={
            "group": incident_group,
            "device_name": None if incident_device_name == "__all__" else incident_device_name,
            "severity": severity,
            "incident_type": incident_type,
        },
    )
    db.commit()
    return RedirectResponse(
        url=f"/incidents?date={report_date}&group={page_group}&incident_message=Инцидент+добавлен",
        status_code=303,
    )


@app.post("/admin/schedules/create")
def create_report_schedule_from_admin(
    request: Request,
    email: str = Form(...),
    group: str = Form("all"),
    send_time: str = Form("08:00"),
    export_format: str = Form("both"),
    template_mode: str = Form("detailed"),
    db: Session = Depends(get_db),
):
    current_user = require_admin(request, db)
    if isinstance(current_user, RedirectResponse):
        return current_user

    if not EMAIL_PATTERN.match(email.strip()):
        return RedirectResponse(url="/admin/users?tab=schedules&schedule_message=Некорректный+email", status_code=303)
    if not re.match(r"^\d{2}:\d{2}$", send_time):
        return RedirectResponse(url="/admin/users?tab=schedules&schedule_message=Некорректный+формат+времени", status_code=303)
    if export_format not in {"pdf", "xlsx", "both"}:
        return RedirectResponse(url="/admin/users?tab=schedules&schedule_message=Неподдерживаемый+формат", status_code=303)
    if template_mode not in {"brief", "detailed"}:
        return RedirectResponse(url="/admin/users?tab=schedules&schedule_message=Неподдерживаемый+шаблон", status_code=303)

    schedule = ReportSchedule(
        recipient_email=email.strip(),
        group_id=(group or "all").strip() or "all",
        export_format=export_format,
        template_mode=template_mode,
        send_time=send_time,
        is_enabled=True,
        created_by_user_id=current_user.id,
    )
    db.add(schedule)
    db.flush()
    _log_audit(
        db,
        actor_user_id=current_user.id,
        action="schedule_created",
        entity_type="report_schedule",
        entity_id=str(schedule.id),
        details={
            "email": schedule.recipient_email,
            "group": schedule.group_id,
            "format": schedule.export_format,
            "time": schedule.send_time,
            "template": schedule.template_mode,
            "source": "admin_panel",
        },
    )
    db.commit()
    return RedirectResponse(url="/admin/users?tab=schedules&schedule_message=Рассылка+создана", status_code=303)


@app.get("/admin/users", response_class=HTMLResponse)
def admin_users(request: Request, db: Session = Depends(get_db)):
    current_user = require_admin(request, db)
    if isinstance(current_user, RedirectResponse):
        return current_user

    log_status = (request.query_params.get("log_status", "all") or "all").strip().lower()

    users = list(db.scalars(select(AppUser).order_by(AppUser.created_at.desc())).all())
    invite_codes = list(db.scalars(select(AdminInviteCode).order_by(AdminInviteCode.created_at.desc()).limit(20)).all())
    schedules = list(db.scalars(select(ReportSchedule).order_by(ReportSchedule.created_at.desc()).limit(20)).all())
    email_logs_stmt = select(EmailReportLog)
    if log_status != "all":
        email_logs_stmt = email_logs_stmt.where(EmailReportLog.status == log_status)
    email_logs = list(db.scalars(email_logs_stmt.order_by(EmailReportLog.created_at.desc()).limit(30)).all())
    queue_items = list(db.scalars(select(ReportDispatchQueue).order_by(ReportDispatchQueue.created_at.desc()).limit(20)).all())
    audit_logs = list(db.scalars(select(AuditLog).order_by(AuditLog.created_at.desc()).limit(250)).all())
    available_groups = EnergyConsumptionService(db).get_available_groups()
    return _render_react_page(
        request,
        page="admin_users",
        title="Админ-панель",
        props=_serialize_admin_payload(
            users=users,
            invite_codes=invite_codes,
            schedules=schedules,
            email_logs=email_logs,
            queue_items=queue_items,
            audit_logs=audit_logs,
            available_groups=available_groups,
            selected_log_status=log_status,
            generated_code=request.query_params.get("generated_code", ""),
            schedule_message=request.query_params.get("schedule_message", ""),
            settings_message=request.query_params.get("settings_message", ""),
            timezone_offset_hours=get_timezone_offset_hours(),
            current_user=current_user,
        ),
        style_paths=["styles/activity.css"],
    )


@app.post("/admin/settings/timezone")
def update_timezone_setting(
    request: Request,
    timezone_offset_hours: int = Form(...),
    db: Session = Depends(get_db),
):
    current_user = require_admin(request, db)
    if isinstance(current_user, RedirectResponse):
        return current_user

    try:
        parsed_offset = _parse_timezone_offset(timezone_offset_hours)
    except ValueError:
        return RedirectResponse(
            url="/admin/users?tab=settings&settings_message=Некорректное+смещение+часового+пояса",
            status_code=303,
        )

    setting = db.get(AppSetting, TIMEZONE_SETTING_KEY)
    if setting is None:
        setting = AppSetting(key=TIMEZONE_SETTING_KEY, value=str(parsed_offset))
        db.add(setting)
    else:
        setting.value = str(parsed_offset)
    set_timezone_offset_hours(parsed_offset)
    request.app.state.timezone_offset_hours = parsed_offset
    _log_audit(
        db,
        actor_user_id=current_user.id,
        action="timezone_updated",
        entity_type="app_settings",
        entity_id=TIMEZONE_SETTING_KEY,
        details={"timezone_offset_hours": parsed_offset},
    )
    db.commit()
    return RedirectResponse(
        url=f"/admin/users?tab=settings&settings_message=Часовой+пояс+обновлен:+UTC{parsed_offset:+d}",
        status_code=303,
    )


@app.post("/admin/users/{user_id}/role")
def update_user_role(
    user_id: int,
    request: Request,
    role: str = Form(...),
    db: Session = Depends(get_db),
):
    current_user = require_admin(request, db)
    if isinstance(current_user, RedirectResponse):
        return current_user

    user = db.get(AppUser, user_id)
    if user is None:
        return RedirectResponse(url="/admin/users", status_code=303)
    if user.id == current_user.id and role != "admin":
        return RedirectResponse(url="/admin/users?settings_message=Нельзя+снять+роль+admin+у+себя", status_code=303)
    if role not in ROLE_CHOICES:
        return RedirectResponse(url="/admin/users?settings_message=Некорректная+роль", status_code=303)

    previous_role = user.role
    user.role = role
    _log_audit(
        db,
        actor_user_id=current_user.id,
        action="user_role_updated",
        entity_type="app_user",
        entity_id=str(user.id),
        details={"username": user.username, "from": previous_role, "to": role},
    )
    db.commit()
    return RedirectResponse(url="/admin/users?settings_message=Роль+обновлена", status_code=303)


@app.post("/admin/users/{user_id}/block")
def block_user(user_id: int, request: Request, db: Session = Depends(get_db)):
    current_user = require_admin(request, db)
    if isinstance(current_user, RedirectResponse):
        return current_user

    user = db.get(AppUser, user_id)
    if user is None:
        return RedirectResponse(url="/admin/users", status_code=303)
    if user.id == current_user.id:
        return RedirectResponse(url="/admin/users", status_code=303)

    user.is_blocked = True
    user.blocked_at = now_local_naive()
    _log_audit(
        db,
        actor_user_id=current_user.id,
        action="user_blocked",
        entity_type="app_user",
        entity_id=str(user.id),
        details={"username": user.username},
    )
    db.commit()
    return RedirectResponse(url="/admin/users", status_code=303)


@app.post("/admin/users/{user_id}/unblock")
def unblock_user(user_id: int, request: Request, db: Session = Depends(get_db)):
    current_user = require_admin(request, db)
    if isinstance(current_user, RedirectResponse):
        return current_user

    user = db.get(AppUser, user_id)
    if user is None:
        return RedirectResponse(url="/admin/users", status_code=303)

    user.is_blocked = False
    user.blocked_at = None
    _log_audit(
        db,
        actor_user_id=current_user.id,
        action="user_unblocked",
        entity_type="app_user",
        entity_id=str(user.id),
        details={"username": user.username},
    )
    db.commit()
    return RedirectResponse(url="/admin/users", status_code=303)


@app.post("/admin/schedules/{schedule_id}/toggle")
def toggle_schedule(schedule_id: int, request: Request, db: Session = Depends(get_db)):
    current_user = require_admin(request, db)
    if isinstance(current_user, RedirectResponse):
        return current_user

    schedule = db.get(ReportSchedule, schedule_id)
    if schedule is None:
        return RedirectResponse(url="/admin/users?tab=schedules", status_code=303)

    schedule.is_enabled = not schedule.is_enabled
    _log_audit(
        db,
        actor_user_id=current_user.id,
        action="schedule_toggled",
        entity_type="report_schedule",
        entity_id=str(schedule.id),
        details={"is_enabled": schedule.is_enabled},
    )
    db.commit()
    return RedirectResponse(url="/admin/users?tab=schedules", status_code=303)


@app.post("/admin/queue/{queue_id}/retry-now")
def retry_queue_now(queue_id: int, request: Request, db: Session = Depends(get_db)):
    current_user = require_admin(request, db)
    if isinstance(current_user, RedirectResponse):
        return current_user

    item = db.get(ReportDispatchQueue, queue_id)
    if item is None:
        return RedirectResponse(url="/admin/users?tab=queue", status_code=303)
    if item.status != "failed":
        return RedirectResponse(url="/admin/users?tab=queue", status_code=303)

    item.status = "pending"
    item.attempts = 0
    item.next_attempt_at = now_local_naive()
    item.last_error = None
    _log_audit(
        db,
        actor_user_id=current_user.id,
        action="queue_retry_now",
        entity_type="report_dispatch_queue",
        entity_id=str(item.id),
        details={"recipient": item.recipient_email},
    )
    db.add(
        EmailReportLog(
            queue_item_id=item.id,
            requested_by_user_id=current_user.id,
            recipient_email=item.recipient_email,
            report_date=item.report_date,
            group_id=item.group_id,
            export_format=item.export_format,
            template_mode=item.template_mode,
            status="retry_manual",
        )
    )
    db.commit()
    return RedirectResponse(url="/admin/users?tab=queue", status_code=303)


@app.post("/admin/invite-codes/generate")
def generate_admin_code(request: Request, db: Session = Depends(get_db)):
    current_user = require_admin(request, db)
    if isinstance(current_user, RedirectResponse):
        return current_user

    invite = generate_admin_invite_code(db, created_by_user_id=current_user.id)
    _log_audit(
        db,
        actor_user_id=current_user.id,
        action="admin_invite_generated",
        entity_type="admin_invite_code",
        entity_id=str(invite.id),
        details={"code": invite.code},
    )
    db.commit()
    return RedirectResponse(url=f"/admin/users?generated_code={invite.code}", status_code=303)
